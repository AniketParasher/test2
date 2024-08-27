import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from io import BytesIO
import os
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch

# Function to generate PDF
def generate_pdf_from_excel(excel_data):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    story = []

    # Create a PDF from Excel data
    # Use ReportLab Canvas to handle drawing
    pdf_canvas = canvas.Canvas(buffer, pagesize=letter)

    # Adjust PDF content here if needed, for example:
    pdf_canvas.drawString(100, 750, "ATTENDANCE LIST")
    pdf_canvas.drawString(100, 735, "(PLEASE FILL ALL THE DETAILS IN BLOCK LETTERS)")

    pdf_canvas.save()
    buffer.seek(0)
    return buffer.read()

# Function to process data and update the template
def process_data_and_update_template(df, template_file):
    result = []

    # Identify columns excluding 'Student ID' and those that are entirely empty
    grouping_columns = [col for col in df.columns if col not in ['STUDENT ID'] and df[col].notna().any()]

    # Group by the identified columns and count the number of unique 'Student ID'
    grouped = df.groupby(grouping_columns).agg(student_count=('STUDENT ID', 'nunique')).reset_index()

    # Process 'CLASS' column if it exists and contains non-numeric values
    if 'CLASS' in grouped.columns and grouped['CLASS'].astype(str).str.contains('\D').any():
        grouped['CLASS'] = grouped['CLASS'].astype(str).str.extract('(\d+)')

    # Convert the result to a list of dictionaries
    result = grouped.to_dict(orient='records')

    pdf_files = []
    for school in result:
        wb = load_workbook(template_file)
        ws = wb.active

        # Update "ATTENDANCE LIST" formatting
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "ATTENDANCE LIST" in str(cell.value):
                    cell.font = Font(name='Calibri', size=20, bold=True)
                    break

        # Update "(PLEASE FILL ALL THE DETAILS IN BLOCK LETTERS)" formatting
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "(PLEASE FILL ALL THE DETAILS IN BLOCK LETTERS)" in str(cell.value):
                    cell.font = Font(name='Calibri', size=9)
                    break

        # Update the PROJECT, DISTRICT, BLOCK, SCHOOL, and CLASS fields
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell_value = str(cell.value)
                    if "PROJECT :" in cell_value:
                        project_city = school.get('PROJECT-CITY', '')
                        cell.value = f"PROJECT : {project_city}"
                    elif "DISTRICT :" in cell_value:
                        district = school.get('District', '')
                        cell.value = f"DISTRICT : {district}"
                    elif "BLOCK :" in cell_value:
                        block = school.get('Block', '')
                        cell.value = f"BLOCK : {block}"
                    elif "SCHOOL :" in cell_value:
                        school_name = school.get('SCHOOL NAME', '')
                        cell.value = f"SCHOOL : {school_name}"
                    elif "CLASS :" in cell_value:
                        class_id = school.get('CLASS', '')
                        cell.value = f"CLASS : {class_id}"

        # Filling in student IDs from DataFrame df
        student_ids = df[df['School Code'] == school['School Code']]['STUDENT ID'].tolist()

        # Find the "STUDENT ID" column
        student_id_column = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and "STUDENT ID" in str(cell.value):
                    student_id_column = cell.column
                    start_row = cell.row + 1
                    break
            if student_id_column:
                break

        # Fill the student IDs in the corresponding column
        if student_id_column:
            for idx, student_id in enumerate(student_ids):
                cell = ws.cell(row=start_row + idx, column=student_id_column, value=student_id)
                cell.font = Font(name='Calibri', size=11)

        # Remove rows after the last student ID entry
        last_row = start_row + len(student_ids)
        for row in ws.iter_rows(min_row=last_row, max_col=ws.max_column):
            for cell in row:
                cell.value = None

        # Apply borders only up to the last filled row
        if student_id_column:
            last_row = start_row + len(student_ids) - 1
            thin_border = Border(left=Side(border_style="thin"),
                                 right=Side(border_style="thin"),
                                 top=Side(border_style="thin"),
                                 bottom=Side(border_style="thin"))

            for row in ws.iter_rows(min_row=last_row + 1, max_row=ws.max_row):
                for cell in row:
                    cell.border = Border()  # Clear borders

        # Center align all rows below the "STUDENT ID" row
        for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Identify the row containing "S.NO" or "STUDENT ID"
        start_row = None
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and ("S.NO" in str(cell.value) or "STUDENT ID" in str(cell.value)):
                    start_row = cell.row
                    break

        # If the start_row is found, set the row height from that row onwards
        if start_row:
            for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
                ws.row_dimensions[row[0].row].height = 60

        # Adjust column widths to fit content
        dim_holder = DimensionHolder(worksheet=ws)
        for col in range(ws.min_column, ws.max_column + 1):
            dim = ColumnDimension(ws, min=col, max=col, width=18)
            dim_holder[get_column_letter(col)] = dim
        ws.column_dimensions = dim_holder

        # Set page margins to narrow
        ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.25, bottom=0.25)

        # Set page size to A4
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        # Save the worksheet to an in-memory file
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Generate PDF from the updated Excel file
        pdf_data = generate_pdf_from_excel(excel_buffer)
        pdf_files.append((f'school_{school["School Code"]}.pdf', pdf_data))
    
    return pdf_files

# Streamlit app
st.title("School Data Processor")

uploaded_excel = st.file_uploader("Upload Excel Data File", type="xlsx")
uploaded_template = st.file_uploader("Upload Template File", type="xlsx")

if uploaded_excel and uploaded_template:
    df = pd.read_excel(uploaded_excel)

    if st.button("Generate PDFs"):
        pdf_files = process_data_and_update_template(df, uploaded_template)

        for pdf_filename, pdf_data in pdf_files:
            st.download_button(
                label=f"Download {pdf_filename}",
                data=pdf_data,
                file_name=pdf_filename,
                mime="application/pdf"
            )
